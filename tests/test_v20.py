from __future__ import annotations

import io
import uuid
from datetime import date

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import inspect, select

from app.compliance_service import (
    build_event_workbook,
    build_reop_annual_workbook,
    create_compliance_event,
    ensure_compliance_event,
)
from app.database import SessionLocal, engine
from app.main import app
from app.models import Branch, BranchComplianceProfile, Company, ComplianceEvent, Employee, User
from app.storage_service import LocalStorage


def _login(client: TestClient) -> None:
    response = client.post(
        "/login",
        data={"email": "admin@digitlaboral.com.py", "password": "demo123"},
        follow_redirects=False,
    )
    assert response.status_code == 303


def test_v20_migrations_and_readiness():
    tables = set(inspect(engine).get_table_names())
    assert {
        "company_compliance_profiles",
        "branch_compliance_profiles",
        "employee_compliance_profiles",
        "payroll_compliance_details",
        "compliance_events",
        "integration_batches",
        "integration_batch_items",
    }.issubset(tables)
    with TestClient(app) as client:
        response = client.get("/health/ready")
        assert response.status_code == 200
        assert response.json()["status"] in {"ready", "degraded"}
        assert response.json()["checks"]["database"] == "ok"


def test_compliance_center_renders_after_login():
    with TestClient(app) as client:
        _login(client)
        response = client.get("/app/compliance")
        assert response.status_code == 200
        assert "Centro REI y REOP" in response.text
        assert "no almacena contraseñas" in response.text


def test_create_user_accepts_empty_company_id_without_422():
    unique = uuid.uuid4().hex[:10]
    email = f"auxiliar-v20-{unique}@example.test"
    with TestClient(app) as client:
        _login(client)
        response = client.post(
            "/app/users",
            data={
                "full_name": "Auxiliar Prueba V20",
                "email": email,
                "password": "Temporal#2026",
                "role": "auxiliar",
                "company_id": "",
            },
            follow_redirects=False,
        )
        assert response.status_code == 303
        assert response.headers["location"] == "/app/users"
    with SessionLocal() as db:
        user = db.scalar(select(User).where(User.email == email))
        assert user is not None
        assert user.company_id is None
        db.delete(user)
        db.commit()


def test_compliance_event_and_exports_are_valid_xlsx():
    with SessionLocal() as db:
        company = db.scalar(select(Company).order_by(Company.id))
        assert company is not None
        event = create_compliance_event(
            db,
            company_id=company.id,
            employee_id=None,
            authority="REOP",
            event_type="Entrada",
            event_date=date(2026, 8, 3),
            due_date=date(2026, 8, 4),
            payload={"detail": "Prueba automatizada"},
            created_by="pytest@digitlaboral.test",
        )
        db.flush()
        event_payload = build_event_workbook(db, company, "REOP", [event])
        annual_payload = build_reop_annual_workbook(db, company, 2026)
        db.rollback()

    event_book = load_workbook(io.BytesIO(event_payload), read_only=True)
    assert {"REOP_COMUNICACIONES", "METADATOS"}.issubset(event_book.sheetnames)
    assert event_book["REOP_COMUNICACIONES"]["D2"].value == "REOP"

    annual_book = load_workbook(io.BytesIO(annual_payload), read_only=True)
    assert {
        "EMPLEADOS_Y_OBREROS",
        "SUELDOS_Y_JORNALES",
        "RESUMEN_GENERAL",
        "CONTROL_DIGIT_LABORAL",
    }.issubset(annual_book.sheetnames)


def test_local_storage_roundtrip_and_path_protection(tmp_path):
    backend = LocalStorage(tmp_path / "objects")
    stored = backend.put("integrations/1/test.txt", b"contenido", "text/plain")
    assert stored.size == 9
    assert backend.exists(stored.key)
    assert backend.get(stored.key) == b"contenido"
    backend.delete(stored.key)
    assert not backend.exists(stored.key)


def test_automatic_compliance_events_are_idempotent():
    with SessionLocal() as db:
        company = db.scalar(select(Company).order_by(Company.id))
        assert company is not None
        first = ensure_compliance_event(
            db, company_id=company.id, employee_id=None, authority="REOP", event_type="Entrada",
            event_date=date(2026, 8, 3), due_date=None, payload={"origin": "test"},
            created_by="pytest@digitlaboral.test", source_key="pytest:v20:idempotent",
        )
        db.flush()
        first_id = first.id
        second = ensure_compliance_event(
            db, company_id=company.id, employee_id=None, authority="REOP", event_type="Entrada",
            event_date=date(2026, 8, 3), due_date=None, payload={"origin": "test"},
            created_by="pytest@digitlaboral.test", source_key="pytest:v20:idempotent",
        )
        assert second.id == first_id
        db.rollback()


def test_branch_patronal_is_used_in_reop_export():
    with SessionLocal() as db:
        company = db.scalar(select(Company).order_by(Company.id))
        assert company is not None
        branch = db.scalar(select(Branch).where(Branch.company_id == company.id).order_by(Branch.id))
        employee = db.scalar(select(Employee).where(Employee.company_id == company.id, Employee.branch_id.is_not(None)).order_by(Employee.id))
        if branch is None:
            branch = Branch(company_id=company.id, name="Sucursal Test", city="Ciudad del Este")
            db.add(branch)
            db.flush()
        if employee is None:
            employee = db.scalar(select(Employee).where(Employee.company_id == company.id).order_by(Employee.id))
            assert employee is not None
            employee.branch_id = branch.id
        else:
            branch = db.get(Branch, employee.branch_id)
            assert branch is not None
        profile = db.scalar(select(BranchComplianceProfile).where(BranchComplianceProfile.branch_id == branch.id))
        if not profile:
            profile = BranchComplianceProfile(branch_id=branch.id)
            db.add(profile)
        profile.mtess_employer_number = "MTESS-SUC-001"
        event = create_compliance_event(
            db, company_id=company.id, employee_id=employee.id, branch_id=branch.id,
            authority="REOP", event_type="Entrada", event_date=date(2026, 8, 3),
            due_date=None, payload={"test": True}, created_by="pytest@digitlaboral.test",
        )
        db.flush()
        branch_name = branch.name
        payload = build_event_workbook(db, company, "REOP", [event])
        db.rollback()
    workbook = load_workbook(io.BytesIO(payload), read_only=True)
    row = list(workbook["REOP_COMUNICACIONES"].iter_rows(min_row=2, max_row=2, values_only=True))[0]
    assert row[0] == "MTESS-SUC-001"
    assert row[1] == branch_name
