from __future__ import annotations

import csv
import io
import json
import re
from dataclasses import dataclass, asdict
from datetime import UTC, date, datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

EMPLOYEE_HEADERS = [
    "nombre_completo",
    "cedula",
    "cargo",
    "fecha_ingreso",
    "salario_base",
    "fecha_nacimiento",
    "tipo_contrato",
    "frecuencia_pago",
    "aporta_ips",
    "correo",
    "telefono",
    "direccion",
    "observaciones",
]

REQUIRED_HEADERS = {"nombre_completo", "cedula", "cargo", "fecha_ingreso", "salario_base"}


@dataclass
class ImportRow:
    row_number: int
    data: dict[str, Any]
    errors: list[str]
    warnings: list[str]

    @property
    def valid(self) -> bool:
        return not self.errors


@dataclass
class ImportPreview:
    company_id: int
    studio_id: int
    user_id: int
    created_at: str
    filename: str
    rows: list[ImportRow]

    @property
    def valid_count(self) -> int:
        return sum(item.valid for item in self.rows)

    @property
    def invalid_count(self) -> int:
        return len(self.rows) - self.valid_count

    def to_dict(self) -> dict[str, Any]:
        return {
            "company_id": self.company_id,
            "studio_id": self.studio_id,
            "user_id": self.user_id,
            "created_at": self.created_at,
            "filename": self.filename,
            "rows": [asdict(row) for row in self.rows],
        }

    @classmethod
    def from_dict(cls, payload: dict[str, Any]) -> "ImportPreview":
        return cls(
            company_id=int(payload["company_id"]),
            studio_id=int(payload["studio_id"]),
            user_id=int(payload["user_id"]),
            created_at=str(payload["created_at"]),
            filename=str(payload.get("filename", "archivo")),
            rows=[ImportRow(**row) for row in payload.get("rows", [])],
        )


def normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"[áàäâ]", "a", text)
    text = re.sub(r"[éèëê]", "e", text)
    text = re.sub(r"[íìïî]", "i", text)
    text = re.sub(r"[óòöô]", "o", text)
    text = re.sub(r"[úùüû]", "u", text)
    text = text.replace("ñ", "n")
    return re.sub(r"[^a-z0-9]+", "_", text).strip("_")


def parse_date(value: Any) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def parse_money(value: Any) -> int | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return max(0, int(round(value)))
    digits = re.sub(r"[^0-9-]", "", str(value))
    try:
        return max(0, int(digits))
    except (TypeError, ValueError):
        return None


def parse_bool(value: Any, default: bool = True) -> bool:
    if value in (None, ""):
        return default
    return normalize_header(value) in {"si", "s", "1", "true", "verdadero", "x", "aporta"}


def _rows_from_xlsx(content: bytes) -> tuple[list[str], list[list[Any]]]:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    iterator = sheet.iter_rows(values_only=True)
    try:
        headers = [normalize_header(value) for value in next(iterator)]
    except StopIteration:
        return [], []
    return headers, [list(row) for row in iterator]


def _rows_from_csv(content: bytes) -> tuple[list[str], list[list[Any]]]:
    decoded = content.decode("utf-8-sig", errors="replace")
    sample = decoded[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel
        dialect.delimiter = ";"
    reader = csv.reader(io.StringIO(decoded), dialect)
    try:
        headers = [normalize_header(value) for value in next(reader)]
    except StopIteration:
        return [], []
    return headers, [list(row) for row in reader]


def parse_employee_import(content: bytes, filename: str, max_rows: int = 2000) -> list[ImportRow]:
    lower = filename.lower()
    if lower.endswith(".xlsx"):
        headers, raw_rows = _rows_from_xlsx(content)
    elif lower.endswith(".csv"):
        headers, raw_rows = _rows_from_csv(content)
    else:
        raise ValueError("El archivo debe ser .xlsx o .csv")

    missing = REQUIRED_HEADERS.difference(headers)
    if missing:
        raise ValueError("Faltan columnas obligatorias: " + ", ".join(sorted(missing)))
    if len(raw_rows) > max_rows:
        raise ValueError(f"El archivo supera el máximo permitido de {max_rows} filas.")

    parsed: list[ImportRow] = []
    for index, values in enumerate(raw_rows, start=2):
        mapping = {headers[pos]: values[pos] if pos < len(values) else None for pos in range(len(headers))}
        if not any(value not in (None, "") for value in mapping.values()):
            continue
        errors: list[str] = []
        warnings: list[str] = []
        name = str(mapping.get("nombre_completo") or "").strip()
        document = re.sub(r"\s+", "", str(mapping.get("cedula") or "").strip())
        position = str(mapping.get("cargo") or "").strip()
        admission = parse_date(mapping.get("fecha_ingreso"))
        salary = parse_money(mapping.get("salario_base"))
        birth = parse_date(mapping.get("fecha_nacimiento"))
        email = str(mapping.get("correo") or "").strip().lower()

        if not name:
            errors.append("Falta el nombre completo.")
        if not document:
            errors.append("Falta la cédula.")
        if not position:
            errors.append("Falta el cargo.")
        if not admission:
            errors.append("La fecha de ingreso no es válida.")
        if salary is None:
            errors.append("El salario base no es válido.")
        if email and "@" not in email:
            warnings.append("El correo parece incompleto.")
        if birth and admission and birth >= admission:
            warnings.append("La fecha de nacimiento debe revisarse.")

        data = {
            "full_name": name,
            "document_number": document,
            "position": position,
            "admission_date": admission.isoformat() if admission else "",
            "base_salary": salary or 0,
            "birth_date": birth.isoformat() if birth else "",
            "contract_type": str(mapping.get("tipo_contrato") or "Tiempo indefinido").strip(),
            "payment_frequency": str(mapping.get("frecuencia_pago") or "Mensual").strip(),
            "ips_contributor": parse_bool(mapping.get("aporta_ips"), default=True),
            "email": email,
            "phone": str(mapping.get("telefono") or "").strip(),
            "address": str(mapping.get("direccion") or "").strip(),
            "notes": str(mapping.get("observaciones") or "").strip(),
        }
        parsed.append(ImportRow(index, data, errors, warnings))
    return parsed


def build_employee_template() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Funcionarios"
    sheet.append(EMPLOYEE_HEADERS)
    sheet.append([
        "Juan Pérez",
        "4567890",
        "Auxiliar administrativo",
        "2026-08-01",
        3044000,
        "2000-05-15",
        "Tiempo indefinido",
        "Mensual",
        "Sí",
        "juan@empresa.com.py",
        "0981000000",
        "Ciudad del Este",
        "Fila de ejemplo: eliminá o reemplazá antes de importar",
    ])
    header_fill = PatternFill("solid", fgColor="173B86")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    widths = [28, 16, 24, 17, 17, 18, 23, 20, 14, 28, 18, 32, 42]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    sheet.freeze_panes = "A2"
    guide = workbook.create_sheet("Guía")
    guide.append(["Columna", "Obligatoria", "Ejemplo / regla"])
    guide_rows = [
        ("nombre_completo", "Sí", "Nombre y apellido"),
        ("cedula", "Sí", "Sin puntos; el sistema conserva guiones"),
        ("cargo", "Sí", "Cargo actual"),
        ("fecha_ingreso", "Sí", "AAAA-MM-DD o DD/MM/AAAA"),
        ("salario_base", "Sí", "Solo número; sin Gs."),
        ("fecha_nacimiento", "No", "AAAA-MM-DD o DD/MM/AAAA"),
        ("aporta_ips", "No", "Sí/No; por defecto Sí"),
    ]
    for row in guide_rows:
        guide.append(row)
    for cell in guide[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
    guide.column_dimensions["A"].width = 26
    guide.column_dimensions["B"].width = 16
    guide.column_dimensions["C"].width = 48
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def save_preview(directory: Path, token: str, preview: ImportPreview) -> Path:
    directory.mkdir(parents=True, exist_ok=True)
    path = directory / f"{token}.json"
    path.write_text(json.dumps(preview.to_dict(), ensure_ascii=False, indent=2), encoding="utf-8")
    return path


def load_preview(directory: Path, token: str, max_age_hours: int = 4) -> ImportPreview:
    if not re.fullmatch(r"[a-f0-9]{32}", token):
        raise ValueError("Token de importación inválido.")
    path = directory / f"{token}.json"
    if not path.exists():
        raise FileNotFoundError("La vista previa ya no está disponible.")
    preview = ImportPreview.from_dict(json.loads(path.read_text(encoding="utf-8")))
    created = datetime.fromisoformat(preview.created_at)
    if created.tzinfo is None:
        created = created.replace(tzinfo=UTC)
    if created < datetime.now(UTC) - timedelta(hours=max_age_hours):
        path.unlink(missing_ok=True)
        raise ValueError("La vista previa expiró. Volvé a subir el archivo.")
    return preview
