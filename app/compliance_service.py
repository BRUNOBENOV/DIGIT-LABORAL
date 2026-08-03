from __future__ import annotations

import hashlib
import io
import json
from collections import Counter
from datetime import date, datetime
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import or_, select
from sqlalchemy.orm import Session

from .models import (
    Branch,
    BranchComplianceProfile,
    Company,
    CompanyComplianceProfile,
    ComplianceEvent,
    Employee,
    EmployeeComplianceProfile,
    IntegrationBatch,
    IntegrationBatchItem,
    Payroll,
    PayrollComplianceDetail,
    PayrollLine,
)

OFFICIAL_SOURCES = {
    "REI": "https://portal.ips.gov.py/sistemas/ipsportal/contenido.php?c=119",
    "REOP": "https://www.mtess.gov.py/?page_id=24021",
}

REOP_EVENT_TYPES = {
    "Entrada", "Salida", "Permiso", "Vacaciones", "Amonestación", "Ausencia",
    "Suspensión", "Preaviso", "Accidente laboral", "Liquidación salarial", "Aguinaldo",
}
REI_EVENT_TYPES = {"Entrada", "Salida", "Declaración salarial", "Aportes", "Reposo"}

COMPLIANCE_STATUS_TRANSITIONS: dict[str, set[str]] = {
    "Borrador": {"Borrador", "Validado", "Anulado"},
    "Validado": {"Validado", "Borrador", "Anulado"},
    "Exportado": {"Exportado", "Presentado", "Observado", "Anulado"},
    "Presentado": {"Presentado", "Aceptado", "Observado", "Rechazado"},
    "Aceptado": {"Aceptado"},
    "Observado": {"Observado", "Borrador", "Validado", "Anulado"},
    "Rechazado": {"Rechazado", "Borrador", "Anulado"},
    "Anulado": {"Anulado"},
}


def can_transition_compliance_status(current: str, target: str) -> bool:
    return target in COMPLIANCE_STATUS_TRANSITIONS.get(current, {current})



def _sheet_style(sheet) -> None:  # noqa: ANN001
    fill = PatternFill("solid", fgColor="173B86")
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.freeze_panes = "A2"
    for column in sheet.columns:
        width = min(42, max(12, max(len(str(cell.value or "")) for cell in column) + 2))
        sheet.column_dimensions[column[0].column_letter].width = width


def _profile_map(db: Session, company_id: int) -> tuple[CompanyComplianceProfile | None, dict[int, EmployeeComplianceProfile]]:
    company_profile = db.scalar(select(CompanyComplianceProfile).where(CompanyComplianceProfile.company_id == company_id))
    employee_profiles = list(
        db.scalars(
            select(EmployeeComplianceProfile)
            .join(Employee, Employee.id == EmployeeComplianceProfile.employee_id)
            .where(Employee.company_id == company_id)
        )
    )
    return company_profile, {item.employee_id: item for item in employee_profiles}


def _branch_context(
    db: Session, company_id: int
) -> tuple[dict[int, Branch], dict[int, BranchComplianceProfile]]:
    branches = list(db.scalars(select(Branch).where(Branch.company_id == company_id)))
    branch_map = {item.id: item for item in branches}
    profiles = list(
        db.scalars(
            select(BranchComplianceProfile)
            .join(Branch, Branch.id == BranchComplianceProfile.branch_id)
            .where(Branch.company_id == company_id)
        )
    )
    return branch_map, {item.branch_id: item for item in profiles}


def _patronal_for_branch(
    company: Company,
    authority: str,
    branch_id: int | None,
    branch_profiles: dict[int, BranchComplianceProfile],
) -> str:
    profile = branch_profiles.get(branch_id) if branch_id else None
    if authority == "REI":
        return (profile.ips_employer_number if profile else "") or company.ips_employer_number
    return (profile.mtess_employer_number if profile else "") or company.mtess_employer_number


def validate_company_for_authority(db: Session, company: Company, authority: str) -> list[str]:
    authority = authority.upper()
    errors: list[str] = []
    company_profile, employee_profiles = _profile_map(db, company.id)
    branch_map, branch_profiles = _branch_context(db, company.id)
    if authority == "REI" and not company.ips_employer_number:
        errors.append("Falta el número patronal IPS de la casa matriz.")
    if authority == "REOP" and not company.mtess_employer_number:
        errors.append("Falta el número patronal MTESS de la casa matriz.")
    if not company_profile:
        errors.append("Falta completar el perfil oficial de la empresa.")
    elif not company_profile.department or not company_profile.district or not company_profile.economic_activity:
        errors.append("El perfil oficial de la empresa está incompleto.")
    employees = list(db.scalars(select(Employee).where(Employee.company_id == company.id, Employee.status == "Activo")))
    checked_branches: set[int] = set()
    for employee in employees:
        profile = employee_profiles.get(employee.id)
        missing = []
        if not employee.document_number:
            missing.append("cédula")
        if not profile:
            missing.append("perfil oficial")
        else:
            for label, value in (("sexo", profile.sex), ("nacionalidad", profile.nationality), ("estado civil", profile.marital_status), ("ocupación", profile.occupation_code)):
                if not value:
                    missing.append(label)
        if employee.branch_id and employee.branch_id in branch_map:
            checked_branches.add(employee.branch_id)
            branch_profile = branch_profiles.get(employee.branch_id)
            patronal = _patronal_for_branch(company, authority, employee.branch_id, branch_profiles)
            if not branch_profile:
                missing.append(f"perfil oficial de sucursal ({branch_map[employee.branch_id].name})")
            if not patronal:
                label = "IPS" if authority == "REI" else "MTESS"
                missing.append(f"número patronal {label} de la sucursal")
        if missing:
            errors.append(f"{employee.full_name}: falta {', '.join(missing)}.")
    for branch_id in checked_branches:
        branch_profile = branch_profiles.get(branch_id)
        if branch_profile and (not branch_profile.department or not branch_profile.district):
            errors.append(f"Sucursal {branch_map[branch_id].name}: ubicación oficial incompleta.")
    return errors


def create_compliance_event(
    db: Session,
    *,
    company_id: int,
    employee_id: int | None,
    branch_id: int | None = None,
    authority: str,
    event_type: str,
    event_date: date,
    due_date: date | None,
    payload: dict,
    created_by: str,
    source_key: str | None = None,
) -> ComplianceEvent:
    authority = authority.upper()
    valid = REI_EVENT_TYPES if authority == "REI" else REOP_EVENT_TYPES
    if event_type not in valid:
        raise ValueError("Tipo de comunicación no reconocido.")
    event = ComplianceEvent(
        company_id=company_id,
        employee_id=employee_id,
        branch_id=branch_id,
        authority=authority,
        event_type=event_type,
        event_date=event_date,
        due_date=due_date,
        status="Borrador",
        source_key=source_key or None,
        payload_json=json.dumps(payload, ensure_ascii=False),
        created_by=created_by,
    )
    db.add(event)
    return event




def ensure_compliance_event(
    db: Session,
    *,
    company_id: int,
    employee_id: int | None,
    branch_id: int | None = None,
    authority: str,
    event_type: str,
    event_date: date,
    due_date: date | None,
    payload: dict,
    created_by: str,
    source_key: str,
) -> ComplianceEvent:
    """Create one draft per source object, avoiding duplicate official communications."""
    source_key = source_key.strip()
    if not source_key:
        raise ValueError("source_key es obligatorio para comunicaciones automáticas.")
    existing = db.scalar(
        select(ComplianceEvent).where(
            ComplianceEvent.company_id == company_id,
            ComplianceEvent.authority == authority.upper(),
            or_(
                ComplianceEvent.source_key == source_key,
                ComplianceEvent.payload_json.contains(f'"source_key": "{source_key}"'),
            ),
            ComplianceEvent.status != "Anulado",
        )
    )
    if existing:
        if not existing.source_key:
            existing.source_key = source_key
        return existing
    merged_payload = {**payload, "source_key": source_key, "generated_automatically": True}
    return create_compliance_event(
        db,
        company_id=company_id,
        employee_id=employee_id,
        branch_id=branch_id,
        authority=authority,
        event_type=event_type,
        event_date=event_date,
        due_date=due_date,
        payload=merged_payload,
        created_by=created_by,
        source_key=source_key,
    )


def build_event_workbook(db: Session, company: Company, authority: str, events: Iterable[ComplianceEvent]) -> bytes:
    authority = authority.upper()
    wb = Workbook()
    ws = wb.active
    ws.title = f"{authority}_COMUNICACIONES"
    ws.append([
        "Nro Patronal", "Establecimiento", "RUC", "Sistema", "Tipo de comunicación", "Cédula", "Trabajador",
        "Fecha del evento", "Vencimiento", "Estado", "Datos complementarios", "ID Digit Laboral",
    ])
    employee_map = {item.id: item for item in db.scalars(select(Employee).where(Employee.company_id == company.id))}
    branch_map, branch_profiles = _branch_context(db, company.id)
    for event in events:
        employee = employee_map.get(event.employee_id) if event.employee_id else None
        branch_id = event.branch_id or (employee.branch_id if employee else None)
        branch = branch_map.get(branch_id) if branch_id else None
        patronal = _patronal_for_branch(company, authority, branch_id, branch_profiles)
        ws.append([
            patronal,
            branch.name if branch else "Casa matriz",
            company.ruc,
            authority,
            event.event_type,
            employee.document_number if employee else "",
            employee.full_name if employee else "",
            event.event_date.isoformat(),
            event.due_date.isoformat() if event.due_date else "",
            event.status,
            event.payload_json,
            event.id,
        ])
    _sheet_style(ws)
    meta = wb.create_sheet("METADATOS")
    meta.append(["Campo", "Valor"])
    meta.append(["Empresa", company.legal_name])
    meta.append(["RUC", company.ruc])
    meta.append(["Sistema", authority])
    meta.append(["Fuente oficial", OFFICIAL_SOURCES[authority]])
    meta.append(["Advertencia", "Archivo de interoperabilidad preparado por Digit Laboral. Verificar contra el modelo oficial vigente antes de presentarlo."])
    _sheet_style(meta)
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def build_reop_annual_workbook(db: Session, company: Company, year: int) -> bytes:
    company_profile, employee_profiles = _profile_map(db, company.id)
    branch_map, branch_profiles = _branch_context(db, company.id)
    employees = list(db.scalars(select(Employee).where(Employee.company_id == company.id).order_by(Employee.full_name)))
    payrolls = list(db.scalars(select(Payroll).where(Payroll.company_id == company.id, Payroll.period.like(f"{year}-%"))))
    payroll_ids = [item.id for item in payrolls]
    lines = list(db.scalars(select(PayrollLine).where(PayrollLine.payroll_id.in_(payroll_ids)))) if payroll_ids else []
    line_details = list(db.scalars(select(PayrollComplianceDetail).where(PayrollComplianceDetail.payroll_line_id.in_([line.id for line in lines])))) if lines else []
    details_by_line = {item.payroll_line_id: item for item in line_details}
    payroll_by_id = {item.id: item for item in payrolls}
    employee_by_id = {item.id: item for item in employees}

    wb = Workbook()
    workers = wb.active
    workers.title = "EMPLEADOS_Y_OBREROS"
    workers.append([
        "Nro Patronal MTESS", "Establecimiento", "RUC", "Tipo Documento", "Documento", "Nombres y apellidos",
        "Nacionalidad", "Estado civil", "Fecha nacimiento", "Sexo", "Departamento", "Distrito",
        "Domicilio", "Profesión/Ocupación", "Código ocupación", "Cargo", "Fecha entrada", "Fecha salida",
    ])
    for employee in employees:
        profile = employee_profiles.get(employee.id)
        branch = branch_map.get(employee.branch_id) if employee.branch_id else None
        workers.append([
            _patronal_for_branch(company, "REOP", employee.branch_id, branch_profiles),
            branch.name if branch else "Casa matriz",
            company.ruc,
            profile.document_type if profile else "CI",
            employee.document_number,
            employee.full_name,
            profile.nationality if profile else "",
            profile.marital_status if profile else "",
            employee.birth_date.isoformat() if employee.birth_date else "",
            profile.sex if profile else "",
            profile.department if profile else "",
            profile.district if profile else "",
            employee.address,
            profile.profession if profile else "",
            profile.occupation_code if profile else "",
            employee.position,
            employee.admission_date.isoformat(),
            employee.termination_date.isoformat() if employee.termination_date else "",
        ])
    _sheet_style(workers)

    wages = wb.create_sheet("SUELDOS_Y_JORNALES")
    wages.append([
        "Nro Patronal MTESS", "Establecimiento", "Periodo", "Documento", "Trabajador", "Desde", "Hasta", "Forma de pago",
        "Días trabajados", "Horas trabajadas", "Salario básico", "Recargo nocturno", "Horas extra diurnas",
        "Horas extra nocturnas", "Feriados", "Vacaciones", "Asignación familiar", "Comisiones",
        "Gratificaciones/Premios", "Bonificaciones", "Otros ingresos", "Aporte IPS trabajador",
        "Otros descuentos", "Total bruto", "Total descuentos", "Neto",
    ])
    for line in lines:
        payroll = payroll_by_id[line.payroll_id]
        employee = employee_by_id.get(line.employee_id)
        detail = details_by_line.get(line.id)
        month_start = f"{payroll.period}-01"
        branch = branch_map.get(employee.branch_id) if employee and employee.branch_id else None
        wages.append([
            _patronal_for_branch(company, "REOP", employee.branch_id if employee else None, branch_profiles),
            branch.name if branch else "Casa matriz",
            payroll.period,
            employee.document_number if employee else "",
            employee.full_name if employee else "",
            detail.period_start.isoformat() if detail and detail.period_start else month_start,
            detail.period_end.isoformat() if detail and detail.period_end else "",
            detail.payment_method if detail else "Transferencia",
            detail.days_worked if detail else 30,
            detail.hours_worked if detail else 0,
            line.base_salary,
            detail.night_surcharge if detail else 0,
            detail.overtime_day if detail else line.overtime,
            detail.overtime_night if detail else 0,
            detail.holidays if detail else 0,
            detail.vacation_pay if detail else 0,
            detail.family_allowance if detail else 0,
            line.commissions,
            line.bonuses,
            line.other_income,
            detail.other_income_detail if detail else 0,
            line.ips_employee,
            line.other_discount + line.advances + line.absences_discount,
            line.gross,
            line.total_discounts,
            line.net,
        ])
    _sheet_style(wages)

    summary = wb.create_sheet("RESUMEN_GENERAL")
    active = [item for item in employees if item.status == "Activo"]
    sex_counts = Counter((employee_profiles.get(item.id).sex if employee_profiles.get(item.id) else "Sin dato") for item in active)
    summary.append(["Concepto", "Cantidad"])
    summary.append(["Total personas activas", len(active)])
    for key, value in sorted(sex_counts.items()):
        summary.append([f"Sexo: {key}", value])
    summary.append(["Empresa", company.legal_name])
    summary.append(["Actividad económica", company_profile.economic_activity if company_profile else ""])
    summary.append(["Año", year])
    _sheet_style(summary)

    metadata = wb.create_sheet("CONTROL_DIGIT_LABORAL")
    metadata.append(["Campo", "Valor"])
    metadata.append(["Generado", datetime.now().isoformat(timespec="seconds")])
    metadata.append(["Fuente oficial de referencia", OFFICIAL_SOURCES["REOP"]])
    metadata.append(["Estado", "Borrador para validación y presentación manual en REOP"])
    metadata.append(["Advertencia", "Antes de presentar, comparar las columnas con el modelo Excel oficial vigente del MTESS."])
    _sheet_style(metadata)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def register_batch(
    db: Session,
    *,
    company: Company,
    authority: str,
    batch_type: str,
    period: str,
    content: bytes,
    created_by: str,
    events: list[ComplianceEvent] | None = None,
) -> IntegrationBatch:
    digest = hashlib.sha256(content).hexdigest()
    batch = IntegrationBatch(
        studio_id=company.studio_id,
        company_id=company.id,
        authority=authority,
        batch_type=batch_type,
        period=period,
        status="Generado",
        file_name=f"{authority.lower()}_{company.ruc}_{period or date.today().isoformat()}.xlsx".replace("/", "-"),
        file_sha256=digest,
        item_count=len(events or []),
        created_by=created_by,
    )
    db.add(batch)
    db.flush()
    for event in events or []:
        db.add(IntegrationBatchItem(batch_id=batch.id, compliance_event_id=event.id, status="Incluido"))
        event.status = "Exportado"
        event.batch_id = batch.id
    return batch
